"""
Common handler - sozlamalar, izohlar, umumiy handlerlar
"""
import logging

from aiogram import Router, F, Bot
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy import select

from database.db import get_session
from database.models import User, TaskComment, TaskStatus
from keyboards.inline import (
    settings_keyboard, back_to_menu_keyboard, cancel_keyboard,
    main_menu_keyboard, language_keyboard, task_list_keyboard,
)
from services.task_service import TaskService
from services.notification_service import NotificationService
from utils.helpers import format_task_detailed
from i18n import t

router = Router()
logger = logging.getLogger(__name__)


class CommentStates(StatesGroup):
    """Izoh qoldirish bosqichlari"""
    waiting_comment = State()


# ===== Sozlamalar =====

def _build_settings_text(user: User) -> str:
    """Tilga mos sozlamalar matnini quradi."""
    lang = user.language
    notif_status = t("settings.notif.on", lang) if user.notifications_enabled else t("settings.notif.off", lang)
    return (
        f"{t('settings.title', lang)}\n\n"
        f"{t('settings.field.name', lang, name=user.full_name)}\n"
        f"{t('settings.field.id', lang, id=user.telegram_id)}\n"
        f"{t('settings.field.lang', lang, lang=user.language.upper())}\n"
        f"{t('settings.field.tz', lang, tz=user.timezone)}\n"
        f"{t('settings.field.notif', lang, status=notif_status)}\n\n"
        f"{t('settings.welcome', lang)}"
    )


# ===== Reply keyboard tugmalari — inline menyu bilan bir xil =====

@router.message(F.text.in_({
    "➕ Yangi vazifa", "➕ Новая задача", "➕ New task",
}))
@router.callback_query(F.data == "menu:newtask")
async def kb_newtask(event, state: FSMContext, user: User):
    """Yangi vazifa yaratish — tasks handler ga yo'naltiradi"""
    from handlers.tasks import cmd_new_task
    if isinstance(event, CallbackQuery):
        await event.answer()
        await cmd_new_task(event.message, state, user)
    else:
        await cmd_new_task(event, state, user)


@router.message(F.text.in_({
    "📋 Mening vazifalarim", "📋 Мои задачи", "📋 My tasks",
}))
async def kb_mytasks(message: Message, user: User):
    """Mening vazifalarim — tasks handler ga yo'naltiradi"""
    from handlers.tasks import cmd_my_tasks
    await cmd_my_tasks(message, user)


@router.message(F.text.in_({
    "📊 Statistika", "📊 Статистика", "📊 Statistics",
}))
@router.callback_query(F.data == "menu:stats")
async def kb_stats(event, user: User):
    """Statistika — stats handler ga yo'naltiradi"""
    from handlers.stats import cmd_stats
    if isinstance(event, CallbackQuery):
        await event.answer()
        await cmd_stats(event.message, user)
    else:
        await cmd_stats(event, user)


@router.message(F.text.in_({
    "⏰ Kechikganlar", "⏰ Просроченные", "⏰ Overdue",
}))
@router.callback_query(F.data == "menu:overdue")
async def kb_overdue(event, user: User):
    """Kechikgan vazifalar"""
    from handlers.tasks import cmd_overdue
    if isinstance(event, CallbackQuery):
        await event.answer()
        await cmd_overdue(event.message, user)
    else:
        await cmd_overdue(event, user)


@router.message(F.text.in_({
    "🏢 Jamoalarim", "🏢 Мои компании", "🏢 My Companies",
}))
@router.callback_query(F.data == "menu:companies")
async def kb_companies(event, user: User):
    """Kompaniyalar — company handler ga yo'naltiradi"""
    from handlers.company import cmd_companies
    if isinstance(event, CallbackQuery):
        await event.answer()
        await cmd_companies(event.message, user)
    else:
        await cmd_companies(event, user)


@router.message(F.text.in_({
    "👥 Guruhlarim", "👥 Мои группы", "👥 My Groups",
}))
@router.callback_query(F.data == "menu:groups")
async def kb_groups(event, user: User):
    """Guruhlar — groups handler ga yo'naltiradi"""
    from handlers.groups import cmd_groups
    if isinstance(event, CallbackQuery):
        await event.answer()
        await cmd_groups(event.message, user)
    else:
        await cmd_groups(event, user)


# ===== Sozlamalar =====

@router.message(Command("settings"))
@router.message(F.text.in_({"⚙️ Sozlamalar", "⚙️ Настройки", "⚙️ Settings"}))
@router.callback_query(F.data == "menu:settings")
async def cmd_settings(event, user: User) -> None:
    """Sozlamalar menyusi (har 3 til uchun)"""
    if isinstance(event, CallbackQuery):
        message = event.message
        edit = True
        await event.answer()
    else:
        message = event
        edit = False

    text = _build_settings_text(user)

    if edit:
        try:
            await message.edit_text(text, reply_markup=settings_keyboard(user))
        except Exception:
            await message.answer(text, reply_markup=settings_keyboard(user))
    else:
        await message.answer(text, reply_markup=settings_keyboard(user))


@router.callback_query(F.data == "settings:toggle_notif")
async def callback_toggle_notif(callback: CallbackQuery, user: User) -> None:
    """Bildirishnomalarni yoqish/o'chirish — keyin sozlamalar menyusiga qaytadi."""
    async with get_session() as session:
        result = await session.execute(
            select(User).where(User.id == user.id)
        )
        db_user = result.scalar_one()
        db_user.notifications_enabled = not db_user.notifications_enabled
        user.notifications_enabled = db_user.notifications_enabled

    toast_key = "settings.notif.toggled_on" if user.notifications_enabled else "settings.notif.toggled_off"
    await callback.answer(t(toast_key, user.language))

    # Sozlamalar menyusiga qaytamiz (yangilangan holat bilan)
    await callback.message.edit_text(
        _build_settings_text(user),
        reply_markup=settings_keyboard(user),
    )


@router.callback_query(F.data == "settings:language")
async def callback_settings_language(callback: CallbackQuery, user: User) -> None:
    """Til o'zgartirish menyusi"""
    await callback.message.edit_text(
        t("lang.choose", user.language),
        reply_markup=language_keyboard(back_target="menu:settings"),
    )
    await callback.answer()


@router.callback_query(F.data == "settings:timezone")
async def callback_settings_timezone(callback: CallbackQuery) -> None:
    """Vaqt mintaqasi"""
    await callback.answer(
        "ℹ️ Hozircha bu funksiya ishlab chiqilmoqda.",
        show_alert=True,
    )


@router.callback_query(F.data == "settings:export")
async def callback_settings_export(callback: CallbackQuery, user: User) -> None:
    """Ma'lumotlarni Excel (.xlsx) formatida eksport qiladi.

    Workbook 2 sheetdan iborat:
      • Vazifalar — barcha vazifalar jadvali (status/muhimlik/deadline va h.k.)
      • Umumiy   — foydalanuvchi ma'lumotlari va statistikasi
    """
    from io import BytesIO
    from datetime import datetime
    from aiogram.types import BufferedInputFile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    lang = user.language or "uz"
    await callback.answer(t("settings.export.preparing", lang))

    async with get_session() as session:
        tasks = await TaskService.get_user_tasks(session, user.id, include_completed=True)

    if not tasks:
        await callback.message.answer(t("settings.export.empty", lang))
        # Sozlamalarga qaytamiz
        await callback.message.answer(
            _build_settings_text(user),
            reply_markup=settings_keyboard(user),
        )
        return

    # ===== Workbook qurish =====
    wb = Workbook()

    # ----- Sheet 1: Vazifalar -----
    ws = wb.active
    ws.title = t("settings.export.sheet_tasks", lang)[:31]

    headers = [
        t("settings.export.col.id", lang),
        t("settings.export.col.title", lang),
        t("settings.export.col.desc", lang),
        t("settings.export.col.status", lang),
        t("settings.export.col.priority", lang),
        t("settings.export.col.deadline", lang),
        t("settings.export.col.created", lang),
    ]
    ws.append(headers)

    # Sarlavha stillari
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="6366F1")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Status / Priority — rangli badge fon
    status_color = {
        "new": "DBEAFE", "in_progress": "FEF3C7", "done": "D1FAE5",
        "cancelled": "FEE2E2", "overdue": "FECACA",
    }
    priority_color = {
        "low": "D1FAE5", "medium": "FEF3C7", "high": "FED7AA", "urgent": "FECACA",
    }

    for task in tasks:
        ws.append([
            task.id,
            task.title,
            task.description or "",
            task.status.value,
            task.priority.value,
            task.deadline.strftime("%Y-%m-%d %H:%M") if task.deadline else "",
            task.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # Body row stillari
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Status rang
        st = row[3].value
        if st in status_color:
            row[3].fill = PatternFill("solid", fgColor=status_color[st])
            row[3].alignment = center
        # Priority rang
        pr = row[4].value
        if pr in priority_color:
            row[4].fill = PatternFill("solid", fgColor=priority_color[pr])
            row[4].alignment = center

    # Ustun kengliklari
    widths = [8, 32, 50, 14, 12, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    # ----- Sheet 2: Umumiy -----
    ws2 = wb.create_sheet(t("settings.export.sheet_summary", lang)[:31])
    ws2.append([
        t("settings.export.summary.field", lang),
        t("settings.export.summary.value", lang),
    ])
    for col_idx in (1, 2):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    done_count = sum(1 for x in tasks if x.status.value == "done")
    overdue_count = sum(1 for x in tasks if x.status.value == "overdue")
    active_count = sum(1 for x in tasks if x.status.value in ("new", "in_progress"))

    summary_rows = [
        (t("settings.export.summary.name", lang), user.full_name),
        (t("settings.export.summary.id", lang), str(user.telegram_id)),
        (t("settings.export.summary.lang", lang), user.language.upper()),
        (t("settings.export.summary.tz", lang), user.timezone),
        (t("settings.export.summary.total", lang), len(tasks)),
        (t("settings.export.summary.done", lang), done_count),
        (t("settings.export.summary.active", lang), active_count),
        (t("settings.export.summary.overdue", lang), overdue_count),
        (t("settings.export.summary.exported_at", lang), datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for row in summary_rows:
        ws2.append(row)

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=2):
        for cell in row:
            cell.border = border
        row[0].font = Font(bold=True)

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 32

    # ----- Faylni yuborish -----
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    file = BufferedInputFile(buf.read(), filename=f"taskbot_export_{user.id}.xlsx")
    await callback.message.answer_document(
        document=file,
        caption=t("settings.export.caption", lang, count=len(tasks)),
    )

    # Sozlamalarga qaytamiz
    await callback.message.answer(
        _build_settings_text(user),
        reply_markup=settings_keyboard(user),
    )


# ===== Vazifa izohlari =====

@router.callback_query(F.data.startswith("task_comments:"))
async def callback_task_comments(callback: CallbackQuery, user: User) -> None:
    """Vazifa izohlarini ko'rsatish"""
    task_id = int(callback.data.split(":")[1])
    
    async with get_session() as session:
        task = await TaskService.get_task(session, task_id)
        if not task:
            await callback.answer("❗ Vazifa topilmadi", show_alert=True)
            return
    
    text = f"💬 <b>{task.title}</b> - izohlar\n\n"
    
    if task.comments:
        for comment in task.comments[-10:]:
            date = comment.created_at.strftime("%d.%m %H:%M")
            author = comment.user.full_name if comment.user else "Noma'lum"
            text += f"👤 <b>{author}</b> <i>({date})</i>\n{comment.content}\n\n"
    else:
        text += "<i>Hali izohlar yo'q. Birinchi bo'lib yozing!</i>\n\n"
    
    text += "💡 Yangi izoh qo'shish uchun tugmani bosing:"
    
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    builder = InlineKeyboardBuilder()
    builder.button(text="➕ Izoh qo'shish", callback_data=f"task_comment_add:{task_id}")
    builder.button(text="🔙 Orqaga", callback_data=f"task_view:{task_id}")
    builder.adjust(1)
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data.startswith("task_comment_add:"))
async def callback_task_comment_add(callback: CallbackQuery, state: FSMContext) -> None:
    """Izoh qo'shish - matn so'rash"""
    task_id = int(callback.data.split(":")[1])
    await state.set_state(CommentStates.waiting_comment)
    await state.update_data(task_id=task_id)
    
    await callback.message.edit_text(
        "💬 <b>Izoh yozing:</b>\n\n"
        "<i>Maksimum 1000 belgi</i>",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@router.message(CommentStates.waiting_comment)
async def process_comment(message: Message, state: FSMContext, user: User, bot: Bot) -> None:
    """Izohni qabul qilish va saqlash"""
    if not message.text:
        await message.answer("❗ Iltimos, matn yuboring.")
        return
    
    if len(message.text) > 1000:
        await message.answer("❗ Izoh juda uzun. Maksimum 1000 belgi.")
        return
    
    data = await state.get_data()
    task_id = data.get("task_id")
    
    if not task_id:
        await state.clear()
        await message.answer("❗ Xatolik. Qaytadan urining.", reply_markup=main_menu_keyboard())
        return
    
    async with get_session() as session:
        comment = await TaskService.add_comment(
            session, task_id, user.id, message.text
        )
        
        task = await TaskService.get_task(session, task_id)
        if task:
            try:
                recipient_ids = {task.creator_id}
                for a in task.assignments:
                    recipient_ids.add(a.user_id)
                await NotificationService.notify_new_comment(
                    bot, session, task, user.full_name, message.text,
                    recipient_ids=recipient_ids,
                )
            except Exception as e:
                logger.warning(f"Izoh notification xatosi: {e}")
    
    await state.clear()
    await message.answer(
        "✅ Izoh qo'shildi!",
        reply_markup=back_to_menu_keyboard(),
    )


# ===== Vazifa tarixi =====

@router.callback_query(F.data.startswith("task_history:"))
async def callback_task_history(callback: CallbackQuery) -> None:
    """Vazifa tarixini ko'rsatish"""
    task_id = int(callback.data.split(":")[1])
    
    async with get_session() as session:
        from database.models import TaskHistory
        from sqlalchemy.orm import selectinload
        
        result = await session.execute(
            select(TaskHistory)
            .where(TaskHistory.task_id == task_id)
            .options(selectinload(TaskHistory.user))
            .order_by(TaskHistory.created_at.desc())
            .limit(20)
        )
        history = list(result.scalars().all())
    
    if not history:
        await callback.answer("📝 Tarix bo'sh", show_alert=True)
        return
    
    text = f"📝 <b>Vazifa tarixi</b> (oxirgi 20 ta)\n\n"
    
    action_names = {
        "created": "📝 Yaratildi",
        "status_changed": "🔄 Status o'zgardi",
        "assigned": "👤 Biriktirildi",
        "comment_added": "💬 Izoh qo'shildi",
        "deadline_changed": "⏰ Deadline o'zgardi",
    }
    
    for h in history:
        date = h.created_at.strftime("%d.%m.%Y %H:%M")
        action = action_names.get(h.action, h.action)
        author = h.user.full_name if h.user else "Tizim"
        text += f"{action}\n<i>{date} - {author}</i>\n\n"
    
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    builder = InlineKeyboardBuilder()
    builder.button(text="🔙 Orqaga", callback_data=f"task_view:{task_id}")
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ===== Vazifani tahrirlash =====

@router.callback_query(F.data.startswith("task_edit:"))
async def callback_task_edit(callback: CallbackQuery, user: User) -> None:
    """Vazifani tahrirlash"""
    task_id = int(callback.data.split(":")[1])
    
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    builder = InlineKeyboardBuilder()
    builder.button(text="📝 Nomini o'zgartirish", callback_data=f"edit_title:{task_id}")
    builder.button(text="📄 Tavsifni o'zgartirish", callback_data=f"edit_desc:{task_id}")
    builder.button(text="⏰ Deadline o'zgartirish", callback_data=f"edit_deadline:{task_id}")
    builder.button(text="⚡ Muhimlikni o'zgartirish", callback_data=f"edit_priority:{task_id}")
    builder.button(text="🔙 Orqaga", callback_data=f"task_view:{task_id}")
    builder.adjust(1)
    
    await callback.message.edit_text(
        "✏️ <b>Vazifani tahrirlash</b>\n\nNimani o'zgartirmoqchisiz?",
        reply_markup=builder.as_markup(),
    )
    await callback.answer()


class EditStates(StatesGroup):
    """Tahrirlash holatlari"""
    waiting_new_title = State()
    waiting_new_desc = State()
    waiting_new_deadline = State()


@router.callback_query(F.data.startswith("edit_title:"))
async def callback_edit_title(callback: CallbackQuery, state: FSMContext) -> None:
    """Nom tahrirlash boshlash"""
    task_id = int(callback.data.split(":")[1])
    await state.set_state(EditStates.waiting_new_title)
    await state.update_data(edit_task_id=task_id)
    
    await callback.message.edit_text(
        "📝 <b>Yangi nomni kiriting:</b>",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@router.message(EditStates.waiting_new_title)
async def process_edit_title(message: Message, state: FSMContext, user: User) -> None:
    """Yangi nomni saqlash"""
    if not message.text or len(message.text) < 3:
        await message.answer("❗ Nom kamida 3 belgi bo'lishi kerak.")
        return
    if len(message.text) > 500:
        await message.answer("❗ Nom maksimum 500 belgi.")
        return
    
    data = await state.get_data()
    task_id = data.get("edit_task_id")
    
    async with get_session() as session:
        task = await TaskService.get_task(session, task_id, load_relations=False)
        if task:
            task.title = message.text
            from database.models import TaskHistory
            history = TaskHistory(
                task_id=task_id, user_id=user.id,
                action="title_changed",
                new_value={"title": message.text},
            )
            session.add(history)
    
    await state.clear()
    await message.answer(
        f"✅ Vazifa nomi o'zgartirildi: <b>{message.text}</b>",
        reply_markup=back_to_menu_keyboard(),
    )


@router.callback_query(F.data.startswith("edit_desc:"))
async def callback_edit_desc(callback: CallbackQuery, state: FSMContext) -> None:
    """Tavsif tahrirlash boshlash"""
    task_id = int(callback.data.split(":")[1])
    await state.set_state(EditStates.waiting_new_desc)
    await state.update_data(edit_task_id=task_id)
    
    await callback.message.edit_text(
        "📄 <b>Yangi tavsifni kiriting:</b>\n\n"
        "<i>O'tkazib yuborish uchun /skip</i>",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@router.message(EditStates.waiting_new_desc, Command("skip"))
async def process_edit_desc_skip(message: Message, state: FSMContext, user: User) -> None:
    """Tavsifni o'chirish"""
    data = await state.get_data()
    task_id = data.get("edit_task_id")
    
    async with get_session() as session:
        task = await TaskService.get_task(session, task_id, load_relations=False)
        if task:
            task.description = None
    
    await state.clear()
    await message.answer(
        "✅ Tavsif o'chirildi.",
        reply_markup=back_to_menu_keyboard(),
    )


@router.message(EditStates.waiting_new_desc)
async def process_edit_desc(message: Message, state: FSMContext, user: User) -> None:
    """Yangi tavsifni saqlash"""
    if not message.text:
        await message.answer("❗ Iltimos, matn yuboring.")
        return
    if len(message.text) > 2000:
        await message.answer("❗ Tavsif maksimum 2000 belgi.")
        return
    
    data = await state.get_data()
    task_id = data.get("edit_task_id")
    
    async with get_session() as session:
        task = await TaskService.get_task(session, task_id, load_relations=False)
        if task:
            task.description = message.text
    
    await state.clear()
    await message.answer(
        "✅ Tavsif yangilandi.",
        reply_markup=back_to_menu_keyboard(),
    )


@router.callback_query(F.data.startswith("edit_deadline:"))
async def callback_edit_deadline(callback: CallbackQuery, state: FSMContext) -> None:
    """Deadline tahrirlash"""
    task_id = int(callback.data.split(":")[1])
    await state.set_state(EditStates.waiting_new_deadline)
    await state.update_data(edit_task_id=task_id)
    
    await callback.message.edit_text(
        "⏰ <b>Yangi deadline kiriting:</b>\n\n"
        "<b>Formatlar:</b>\n"
        "• <code>25.04.2026 18:00</code>\n"
        "• <code>25.04.2026</code>\n"
        "• <code>2026-04-25 18:00</code>\n\n"
        "Deadline olib tashlash uchun /skip",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@router.message(EditStates.waiting_new_deadline, Command("skip"))
async def process_edit_deadline_skip(message: Message, state: FSMContext, user: User) -> None:
    """Deadline olib tashlash"""
    data = await state.get_data()
    task_id = data.get("edit_task_id")
    
    async with get_session() as session:
        task = await TaskService.get_task(session, task_id, load_relations=False)
        if task:
            task.deadline = None
    
    await state.clear()
    await message.answer(
        "✅ Deadline olib tashlandi.",
        reply_markup=back_to_menu_keyboard(),
    )


@router.message(EditStates.waiting_new_deadline)
async def process_edit_deadline(message: Message, state: FSMContext, user: User) -> None:
    """Yangi deadline saqlash"""
    from utils.helpers import parse_datetime
    from datetime import datetime
    
    if not message.text:
        await message.answer("❗ Iltimos, sana kiriting.")
        return
    
    deadline = parse_datetime(message.text)
    if not deadline:
        await message.answer(
            "❗ Sana formati noto'g'ri.\n\n"
            "To'g'ri formatlar:\n"
            "• <code>25.04.2026 18:00</code>\n"
            "• <code>25.04.2026</code>"
        )
        return
    
    if deadline < datetime.now():
        await message.answer("❗ Deadline o'tib ketgan sana bo'lmasligi kerak.")
        return
    
    data = await state.get_data()
    task_id = data.get("edit_task_id")
    
    async with get_session() as session:
        task = await TaskService.get_task(session, task_id, load_relations=False)
        if task:
            task.deadline = deadline
            task.warned_24h = False
            task.warned_1h = False
    
    await state.clear()
    await message.answer(
        f"✅ Deadline yangilandi: <b>{deadline.strftime('%d.%m.%Y %H:%M')}</b>",
        reply_markup=back_to_menu_keyboard(),
    )


@router.callback_query(F.data.startswith("edit_priority:"))
async def callback_edit_priority(callback: CallbackQuery) -> None:
    """Muhimlik tahrirlash"""
    task_id = int(callback.data.split(":")[1])
    
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    builder = InlineKeyboardBuilder()
    builder.button(text="🟢 Past", callback_data=f"set_priority:{task_id}:low")
    builder.button(text="🟡 O'rta", callback_data=f"set_priority:{task_id}:medium")
    builder.button(text="🟠 Yuqori", callback_data=f"set_priority:{task_id}:high")
    builder.button(text="🔴 Juda muhim", callback_data=f"set_priority:{task_id}:urgent")
    builder.button(text="🔙 Orqaga", callback_data=f"task_edit:{task_id}")
    builder.adjust(2, 2, 1)
    
    await callback.message.edit_text(
        "⚡ <b>Yangi muhimlik darajasini tanlang:</b>",
        reply_markup=builder.as_markup(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("set_priority:"))
async def callback_set_priority(callback: CallbackQuery, user: User) -> None:
    """Muhimlikni saqlash"""
    parts = callback.data.split(":")
    task_id = int(parts[1])
    new_priority = parts[2]
    
    from database.models import Priority
    
    async with get_session() as session:
        task = await TaskService.get_task(session, task_id, load_relations=False)
        if task:
            task.priority = Priority(new_priority)
    
    priority_names = {"low": "🟢 Past", "medium": "🟡 O'rta", "high": "🟠 Yuqori", "urgent": "🔴 Juda muhim"}
    await callback.message.edit_text(
        f"✅ Muhimlik o'zgartirildi: <b>{priority_names.get(new_priority, new_priority)}</b>",
        reply_markup=back_to_menu_keyboard(),
    )
    await callback.answer()


# ===== Vazifalarni filtrlash =====

@router.callback_query(F.data.startswith("filter:"))
async def callback_filter_tasks(callback: CallbackQuery, user: User) -> None:
    """Vazifalarni statusga ko'ra filtrlash"""
    filter_value = callback.data.split(":")[1]
    
    async with get_session() as session:
        if filter_value == "all":
            tasks = await TaskService.get_user_tasks(session, user.id, include_completed=True)
        else:
            status = TaskStatus(filter_value)
            tasks = await TaskService.get_user_tasks(session, user.id, status=status)
    
    if not tasks:
        status_names = {
            "new": "yangi", "in_progress": "jarayondagi", "review": "ko'rilayotgan",
            "done": "bajarilgan", "overdue": "kechikkan", "all": "barcha",
        }
        name = status_names.get(filter_value, filter_value)
        await callback.message.edit_text(
            f"📋 Hech qanday {name} vazifa topilmadi.",
            reply_markup=back_to_menu_keyboard(),
        )
        await callback.answer()
        return
    
    await callback.message.edit_text(
        f"📋 <b>Filtrlangan vazifalar</b> ({len(tasks)} ta)\n\nVazifani tanlang:",
        reply_markup=task_list_keyboard(tasks),
    )
    await callback.answer()


# ===== Umumiy buyruqlar =====

@router.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext) -> None:
    """Joriy amalni bekor qilish"""
    current_state = await state.get_state()
    if current_state is None:
        await message.answer("ℹ️ Bekor qilinadigan amal yo'q.")
        return
    
    await state.clear()
    await message.answer(
        "❌ Amal bekor qilindi.",
        reply_markup=main_menu_keyboard(),
    )


@router.message(F.text & ~F.text.startswith("/"))
async def fallback_text(message: Message, state: FSMContext) -> None:
    """Noma'lum matnlarga javob"""
    current_state = await state.get_state()
    if current_state is not None:
        return
    
    if message.chat.type in ("group", "supergroup"):
        return
    
    await message.answer(
        "🤔 Buyruqni tushunmadim.\n\n"
        "Yordam uchun /help yoki asosiy menyu uchun /start yuboring.",
        reply_markup=main_menu_keyboard(),
    )
